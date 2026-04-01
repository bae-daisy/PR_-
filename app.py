#!/usr/bin/env python3
"""
PR 데이터 요청 처리 웹앱
- 요청사항 + CSV 업로드 → 자동 파악 → 엑셀 다운로드
- 지표(MAU/DAU/WAU/신규설치), 연령대, 앱 이름 모두 CSV에서 자동 감지
"""
import csv
import re
import io
import base64
from datetime import date
from flask import Flask, request, render_template, send_file, jsonify, make_response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

# ── 스타일 ──
FONT_STYLE = Font(name="맑은 고딕", size=12)
THIN_BORDER = Border(
    top=Side(style="thin"), bottom=Side(style="thin"),
    left=Side(style="thin"), right=Side(style="thin"),
)
HEADER_FILL = PatternFill(patternType="solid", fgColor=Color(theme=3, tint=0.9))
NUM_FMT = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"_);_(@_)'
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_CONT = Alignment(horizontal="centerContinuous", vertical="center")
VCENTER = Alignment(vertical="center")


def apply_style(cell, is_header=False, is_metric=False, is_number=False):
    cell.font = FONT_STYLE
    cell.border = THIN_BORDER
    if is_metric:
        cell.alignment = CENTER_CONT
        cell.fill = HEADER_FILL
    elif is_header:
        cell.alignment = CENTER
    elif is_number:
        cell.alignment = VCENTER
        cell.number_format = NUM_FMT
    else:
        cell.alignment = CENTER


def detect_metric(header_lines):
    text = " ".join(header_lines)
    if "신규 설치" in text or "신규설치" in text:
        label = "신규설치건수"
    elif "사용자 수" in text:
        label = "AU"
    else:
        label = "AU"
    if "월간" in text:
        p = "M"
    elif "주간" in text:
        p = "W"
    elif "일간" in text:
        p = "D"
    else:
        p = "M"
    if label == "AU":
        return f"{p}AU"
    prefixes = {"M": "월간", "W": "주간", "D": "일간"}
    return f"{prefixes[p]} {label}"


def detect_age_group(header_lines):
    text = " ".join(header_lines)
    m = re.search(r"연령:\s*(.+?)\)", text)
    return m.group(1).strip() if m else None


def detect_period_type(dates):
    if not dates:
        return "monthly"
    s = dates[0]
    if re.match(r"^\d{4}-\d{2}$", s):
        return "monthly"
    if "~" in s:
        return "weekly"
    return "daily"


def format_date_label(date_str, period_type):
    if period_type == "monthly":
        m = re.match(r"(\d{4})-(\d{2})", date_str)
        if m:
            return f"{m.group(1)}년 {int(m.group(2))}월"
    return date_str


def parse_csv_content(content):
    lines = content.strip().split("\n")
    header_lines, col_headers, data_rows = [], None, []
    data_started = False
    for line in lines:
        stripped = line.strip().strip('"')
        if not data_started:
            row = list(csv.reader([line]))[0]
            if not row or not row[0].strip():
                header_lines.append(stripped)
                continue
            first = row[0].strip().strip('"')
            if first in ("패키지명", "날짜"):
                data_started = True
                col_headers = [c.strip().strip('"') for c in row]
                continue
            header_lines.append(stripped)
        else:
            row = list(csv.reader([line]))[0]
            if row and len(row) >= 2:
                data_rows.append([c.strip().strip('"') for c in row])
    return header_lines, col_headers, data_rows


def parse_app_names_from_request(request_text):
    """요청사항에서 앱 이름 목록을 순서대로 추출"""
    if not request_text:
        return []
    # "앱:" 또는 "앱 :" 뒤의 내용에서 추출
    m = re.search(r"앱\s*[:：]\s*(.+)", request_text)
    if m:
        line = m.group(1).strip()
        apps = [a.strip() for a in re.split(r"[,，/]", line) if a.strip()]
        return apps
    return []


def match_app_name(csv_name, requested_apps):
    """CSV 앱 이름을 요청사항의 앱 이름과 매칭. 매칭되면 (요청 이름, 순서) 반환"""
    csv_lower = csv_name.lower().replace(" ", "")
    for i, req in enumerate(requested_apps):
        req_lower = req.lower().replace(" ", "")
        # 포함 관계로 매칭 (양방향)
        if req_lower in csv_lower or csv_lower in req_lower:
            return req, i
    # 매칭 안 되면 원본 이름 그대로
    return csv_name, None


def convert(csv_files, media_name, request_text=""):
    """CSV 파일들을 분석하여 자동으로 엑셀 생성. 반환: (BytesIO, warnings, preview, cross_check, app_mappings)"""
    warnings = []
    requested_apps = parse_app_names_from_request(request_text)
    cross_check = []  # 교차 확인 데이터
    app_mappings = []  # CSV앱→표시앱 매칭 정보

    # 1) 모든 CSV 파싱
    parsed = []
    for filename, content in csv_files:
        header_lines, col_headers, data_rows = parse_csv_content(content)
        if not col_headers:
            warnings.append(f"'{filename}' 파일에서 데이터를 찾을 수 없어요. 다시 확인해주세요.")
            continue
        if not data_rows:
            warnings.append(f"'{filename}' 파일에 데이터 행이 없어요. 다시 확인해주세요.")
            continue
        metric = detect_metric(header_lines)
        age = detect_age_group(header_lines)
        first_col = col_headers[0]

        # 날짜/앱 컬럼 추출
        if first_col == "패키지명":
            dates = [row[1] for row in data_rows]
            vals = []
            for row in data_rows:
                raw = row[2].replace(",", "") if len(row) > 2 else ""
                vals.append(int(raw) if raw and raw != "-" else None)
            # 패키지명 형태에서도 요청사항 앱 매칭 시도
            if requested_apps:
                # 헤더에서 앱 이름 힌트 찾기
                pkg_name = data_rows[0][0] if data_rows else ""
                display_name, order = match_app_name(pkg_name, requested_apps)
                if order is None:
                    display_name = requested_apps[0] if len(requested_apps) == 1 else media_name
                app_mappings.append({"csv": pkg_name or media_name, "display": display_name})
            else:
                display_name = media_name
            app_columns = [(display_name, vals)]
            raw_app_columns = [(pkg_name if data_rows else media_name, vals)]
        else:
            dates = [row[0] for row in data_rows]
            app_columns = []
            raw_app_columns = []
            for i in range(1, len(col_headers)):
                csv_name = col_headers[i]
                short = csv_name.split("–")[0].split("-")[0].strip()
                vals = []
                for row in data_rows:
                    raw = row[i].replace(",", "") if i < len(row) else ""
                    vals.append(int(raw) if raw and raw != "-" and raw else None)
                # 요청사항 앱 이름 매칭
                display_name, order = match_app_name(short, requested_apps)
                app_columns.append((display_name, vals, order, short))
                raw_app_columns.append((short, vals))
                app_mappings.append({"csv": short, "display": display_name})

            # 요청사항에 앱 순서가 있으면 그 순서대로 정렬
            if requested_apps and any(a[2] is not None for a in app_columns):
                app_columns.sort(key=lambda a: a[2] if a[2] is not None else 999)

        # app_columns에서 order, original 제거
        clean_apps = [(a[0], a[1]) for a in app_columns] if first_col != "패키지명" else app_columns

        parsed.append({
            "metric": metric,
            "age": age or "전체",
            "dates": dates,
            "apps": clean_apps,
            "_cc": {
                "filename": filename,
                "metric": metric,
                "age": age or "전체",
                "raw_apps": [(a[0], a[1]) for a in raw_app_columns] if first_col != "패키지명" else raw_app_columns,
                "dates": dates,
            },
        })

    if not parsed:
        return None, warnings, [], [], []

    # 2) 그룹핑: (metric, age) 조합별로 정렬
    age_order = ["전체", "10대 이하", "20대", "30대", "40대", "50대", "60대 이상"]
    metric_order = ["MAU", "WAU", "DAU", "월간 신규설치건수", "일간 신규설치건수"]

    def sort_key(item):
        m = item["metric"]
        mi = metric_order.index(m) if m in metric_order else 99
        a = item["age"]
        ai = age_order.index(a) if a in age_order else 99
        return (mi, ai)

    parsed.sort(key=sort_key)
    # cross_check도 같은 순서로
    cross_check = [item["_cc"] for item in parsed]

    # 3) 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = f"MI-{media_name}"

    col_cursor = 2
    has_ages = len(set(p["age"] for p in parsed)) > 1

    for item in parsed:
        dates = item["dates"]
        apps = item["apps"]
        metric = item["metric"]
        age = item["age"]
        period_type = detect_period_type(dates)
        block_width = 1 + len(apps)

        # 지표 라벨
        label = f"{metric} ({age})" if has_ages else metric

        # 6행: 지표 헤더
        for c in range(block_width):
            apply_style(ws.cell(row=6, column=col_cursor + c), is_metric=True)
        ws.cell(row=6, column=col_cursor).value = label

        # 7행: 날짜 + 앱 헤더
        ws.cell(row=7, column=col_cursor).value = "날짜"
        apply_style(ws.cell(row=7, column=col_cursor), is_header=True)
        # 날짜 열 너비: 내용 길이에 맞게 자동 조정
        max_date_len = max((len(format_date_label(d, period_type)) for d in dates), default=6)
        ws.column_dimensions[get_column_letter(col_cursor)].width = max(max_date_len * 1.3 + 2, 12)

        for idx, (app_name, _) in enumerate(apps):
            vc = col_cursor + 1 + idx
            ws.cell(row=7, column=vc).value = app_name
            apply_style(ws.cell(row=7, column=vc), is_header=True)
            app_name_len = max(len(app_name), 7)
            ws.column_dimensions[get_column_letter(vc)].width = max(app_name_len * 1.3 + 2, 13)

        # 8행~: 데이터
        for i, d in enumerate(dates):
            r = 8 + i
            dc = ws.cell(row=r, column=col_cursor)
            dc.value = format_date_label(d, period_type)
            dc.number_format = "@"
            apply_style(dc)
            for idx, (_, vals) in enumerate(apps):
                vc = col_cursor + 1 + idx
                cell = ws.cell(row=r, column=vc)
                if vals[i] is None:
                    cell.value = "-"
                    cell.font = FONT_STYLE
                    cell.border = THIN_BORDER
                    cell.alignment = CENTER
                else:
                    cell.value = vals[i]
                    apply_style(cell, is_number=True)

        col_cursor += block_width + 1

    # 4) 미리보기 데이터 생성
    preview = []
    has_ages = len(set(p["age"] for p in parsed)) > 1
    for item in parsed:
        dates = item["dates"]
        apps = item["apps"]
        metric = item["metric"]
        age = item["age"]
        period_type = detect_period_type(dates)
        label = f"{metric} ({age})" if has_ages else metric
        headers = ["날짜"] + [a[0] for a in apps]
        rows = []
        for i, d in enumerate(dates):
            row = [format_date_label(d, period_type)]
            for _, vals in apps:
                v = vals[i]
                row.append(f"{v:,}" if v is not None else "-")
            rows.append(row)
        preview.append({"label": label, "headers": headers, "rows": rows})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # 중복 제거
    seen = set()
    unique_mappings = []
    for m in app_mappings:
        key = m["csv"] + "→" + m["display"]
        if key not in seen:
            seen.add(key)
            unique_mappings.append(m)
    return buf, warnings, preview, cross_check, unique_mappings


# ── 라우트 ──
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/convert", methods=["POST"])
def api_convert():
    media_name = request.form.get("media_name", "").strip()
    if not media_name:
        return jsonify({"error": "매체명을 입력해주세요."}), 400

    request_text = request.form.get("request_text", "").strip()

    files = request.files.getlist("csv_files")
    if not files or not files[0].filename:
        return jsonify({"error": "CSV 파일을 업로드해주세요."}), 400

    csv_files = []
    for f in files:
        content = f.read().decode("utf-8-sig")
        csv_files.append((f.filename, content))

    buf, warnings, preview, cross_check, app_mappings = convert(csv_files, media_name, request_text)
    if not buf:
        error_msg = "변환에 실패했습니다. CSV 형식을 확인해주세요."
        return jsonify({"error": error_msg, "warnings": warnings}), 400

    filename = f"MI-{media_name}_요청_데이터.xlsx"
    file_b64 = base64.b64encode(buf.read()).decode("ascii")

    # 교차 확인 데이터 직렬화
    cross_check_out = []
    for cc in cross_check:
        apps_data = []
        for app_name, vals in cc["raw_apps"]:
            apps_data.append({
                "name": app_name,
                "values": [f"{v:,}" if v is not None else "-" for v in vals]
            })
        period_type = detect_period_type(cc["dates"])
        cross_check_out.append({
            "filename": cc["filename"],
            "metric": cc["metric"],
            "age": cc["age"],
            "dates": [format_date_label(d, period_type) for d in cc["dates"]],
            "apps": apps_data,
        })

    return jsonify({
        "filename": filename,
        "file": file_b64,
        "warnings": warnings,
        "preview": preview,
        "crossCheck": cross_check_out,
        "appMappings": app_mappings,
    })


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
